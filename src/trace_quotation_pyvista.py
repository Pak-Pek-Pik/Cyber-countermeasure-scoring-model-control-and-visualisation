"""3D rendering (PyVista/VTK) and statistical control of a scoring grid.

Reads an Excel workbook describing countermeasures rated on four criteria
(effectiveness E, versatility P, applicability A, cost C), checks the applied
scale against the documented one, then produces:

  pv_map_model.png   : model plane at median A and C, with the actual
                       countermeasures and their deviation from the plane
  pv_surface_rbf.png : interpolated actual surface (RBF thin-plate spline),
                       masked to the convex hull of the data so that nothing is
                       extrapolated, over the theoretical plane
  pv_map_iso.png     : iso-score map seen from above
  pv_residues.png    : difference between the actual score and the plane

An interactive HTML export (quotation_interactive.html) is attempted at the
end of the script.

The 3D views go through a normalised rendering frame: the raw ranges of E, P
and the score differ enough that the plane would otherwise come out flattened.
Displayed axes stay graduated in real units.

Expected sheet, from row 3 onwards:
    A = countermeasure name, B = status, D = E, E = P, F = A, G = C, H = score.
Column E holds a COUNTIF formula, so the workbook is opened with
data_only=True to read the computed value rather than the formula itself.

Usage:
    python trace_quotation_pyvista.py [--data data/parades.xlsx]
                                      [--output figures]
                                      [--sheet "Parades BdD"]
                                      [--interactive]
"""

from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
import pyvista as pv
from openpyxl import load_workbook
from scipy import stats
from scipy.interpolate import RBFInterpolator
from scipy.linalg import lstsq
from scipy.spatial import Delaunay

parser = argparse.ArgumentParser(
    description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
parser.add_argument("--data", type=Path, default=Path("data/parades.xlsx"),
                    help="source Excel workbook")
parser.add_argument("--output", type=Path, default=Path("figures"),
                    help="destination folder for the PNG files")
parser.add_argument("--sheet", default="Parades BdD",
                    help="name of the sheet to read")
parser.add_argument("--interactive", action="store_true",
                    help="open PyVista windows instead of rendering off-screen")
args = parser.parse_args()

SOURCE, OUTPUT, SHEET = args.data, args.output, args.sheet
OUTPUT.mkdir(parents=True, exist_ok=True)

W_E, W_P, W_A, W_C = 0.35, 0.25, 0.20, 0.20
N_GRID = 1500
SIZE = (1500, 1050)
BOX = (10.0, 10.0, 10.0)

pv.OFF_SCREEN = not args.interactive
pv.set_plot_theme("document")


def read_countermeasures(source: Path, sheet: str) -> pd.DataFrame:
    """Columns A, B and D to H of the sheet, with formulas already evaluated."""
    ws = load_workbook(source, data_only=True)[sheet]
    rows = []
    for r in range(3, ws.max_row + 1):
        E, P, A, C, score = (ws[f"{c}{r}"].value for c in "DEFGH")
        if None in (E, P, A, C, score):
            continue
        rows.append({"name": ws[f"A{r}"].value, "status": ws[f"B{r}"].value,
                     "E": float(E), "P": float(P), "A": float(A),
                     "C": float(C), "score": float(score)})
    return pd.DataFrame(rows)


df = read_countermeasures(SOURCE, SHEET)
print(f"{len(df)} countermeasures read")
print(df[["E", "P", "A", "C", "score"]].describe().round(2), "\n")

A_ref, C_ref = df["A"].median(), df["C"].median()
offset = W_A * A_ref + W_C * C_ref
print(f"Section at A = {A_ref:g} and C = {C_ref:g}  ->  offset = {offset:.2f}")

X = df[["E", "P", "A", "C"]].to_numpy()
y = df["score"].to_numpy()
weights, *_ = lstsq(X, y)
print("Re-estimated weights (lstsq)  E, P, A, C:", np.round(weights, 4))
print("Documented weights            E, P, A, C:", [W_E, W_P, W_A, W_C])

plane_obs = W_E * df["E"] + W_P * df["P"] + offset
r2_plane = 1 - ((y - plane_obs) ** 2).sum() / ((y - y.mean()) ** 2).sum()
df["residual"] = y - plane_obs
print(f"R2 of the (E,P) plane against the actual scores: {r2_plane:.3f}")

for col in ["E", "P", "A", "C"]:
    r, p_r = stats.pearsonr(df[col], df["score"])
    rho, _ = stats.spearmanr(df[col], df["score"])
    print(f"  {col}: Pearson r = {r:+.3f} (p = {p_r:.1e})   Spearman rho = {rho:+.3f}")

t, p_t = stats.ttest_1samp(df["residual"], 0.0)
print(f"Mean residual = {df['residual'].mean():+.3f}  (t = {t:+.2f}, p = {p_t:.3f})")

groups = [g["score"].to_numpy() for _, g in df.groupby("status")]
if len(groups) == 2:
    u, p_u = stats.mannwhitneyu(*groups)
    n1, n2 = sorted(df["status"].unique())
    print(f"Mann-Whitney {n1} vs {n2}: U = {u:.0f}, p = {p_u:.3f}\n")

e_lin = np.linspace(df["E"].min(), df["E"].max(), N_GRID)
p_lin = np.linspace(df["P"].min(), df["P"].max(), N_GRID)
EE, PP = np.meshgrid(e_lin, p_lin)

ZZ_plane = W_E * EE + W_P * PP + offset

agg = (df.groupby(["E", "P"])
         .agg(n=("name", "size"), score=("score", "mean"))
         .reset_index())
rbf = RBFInterpolator(agg[["E", "P"]].to_numpy(), agg["score"].to_numpy(),
                      kernel="thin_plate_spline", smoothing=0.5)
ZZ_rbf = rbf(np.column_stack([EE.ravel(), PP.ravel()])).reshape(EE.shape)

hull = Delaunay(agg[["E", "P"]].to_numpy())
inside = (hull.find_simplex(np.column_stack([EE.ravel(), PP.ravel()])) >= 0)
ZZ_rbf = np.where(inside.reshape(EE.shape), ZZ_rbf, np.nan)


class Frame:
    """Maps (E, P, score) into a roughly cubic rendering box.

    Displayed axes stay graduated in real units through show_grid and
    axes_ranges: the normalisation only prevents a flattened plane.
    """

    def __init__(self, e, p, z, box=BOX):
        self.lo = np.array([np.nanmin(e), np.nanmin(p), np.nanmin(z)])
        self.hi = np.array([np.nanmax(e), np.nanmax(p), np.nanmax(z)])
        self.L = np.asarray(box, float)

    def __call__(self, e, p, z):
        a = (np.asarray(e, float) - self.lo[0]) / (self.hi[0] - self.lo[0]) * self.L[0]
        b = (np.asarray(p, float) - self.lo[1]) / (self.hi[1] - self.lo[1]) * self.L[1]
        c = (np.asarray(z, float) - self.lo[2]) / (self.hi[2] - self.lo[2]) * self.L[2]
        return a, b, c

    @property
    def ranges(self):
        return [self.lo[0], self.hi[0], self.lo[1], self.hi[1],
                self.lo[2], self.hi[2]]


z_all = np.concatenate([ZZ_plane.ravel(), y, ZZ_rbf[~np.isnan(ZZ_rbf)]])
frame = Frame(EE, PP, z_all)


def grid(zz, name="score"):
    """PyVista StructuredGrid built from a 2D array of scores."""
    x, yb, z = frame(EE, PP, zz)
    g = pv.StructuredGrid(x, yb, np.nan_to_num(z, nan=0.0))
    g[name] = zz.ravel(order="F")
    return g


plane_mesh = grid(ZZ_plane)
rbf_mesh = grid(ZZ_rbf)

xn, yn, zn = frame(df["E"], df["P"], df["score"])
cloud = pv.PolyData(np.column_stack([xn, yn, zn]))
cloud["score"] = y
cloud["residual"] = df["residual"].to_numpy()

stems = pv.PolyData()
for e, p_, s in zip(df["E"], df["P"], df["score"]):
    z_plane = W_E * e + W_P * p_ + offset
    a, b, c1 = frame(e, p_, s)
    _, _, c0 = frame(e, p_, z_plane)
    stems += pv.Line((a, b, float(c0)), (a, b, float(c1)))

CB = dict(title="Score", vertical=True, position_x=0.87, position_y=0.2,
          height=0.6, width=0.05, title_font_size=20, label_font_size=16,
          n_labels=6, fmt="%.1f")


def axes_frame(p, ztitle="Score"):
    p.show_grid(xtitle="Effectiveness E", ytitle="Versatility P", ztitle=ztitle,
                axes_ranges=frame.ranges, font_size=11, n_xlabels=5,
                n_ylabels=5, n_zlabels=5, padding=0.0)
    p.view_isometric()
    p.camera.azimuth = -30
    p.camera.elevation = 10
    p.camera.zoom(0.95)


p = pv.Plotter(off_screen=pv.OFF_SCREEN, window_size=SIZE)
p.add_mesh(plane_mesh, scalars="score", cmap="viridis", opacity=0.55,
           show_scalar_bar=False, smooth_shading=True)
p.add_mesh(plane_mesh.contour(12, scalars="score"), color="white",
           line_width=3, render_lines_as_tubes=True)
p.add_mesh(stems, color="dimgrey", line_width=1.5, opacity=0.7)
p.add_mesh(cloud, scalars="score", cmap="plasma", point_size=20,
           render_points_as_spheres=True, scalar_bar_args=CB)
p.add_text(f"Score = 0.35 E + 0.25 P + 0.20 A + 0.20 C\n"
           f"plane at A = {A_ref:g} and C = {C_ref:g}  -  "
           f"{len(df)} countermeasures  -  R2 = {r2_plane:.2f}",
           font_size=13, position="upper_left")
axes_frame(p)
p.show(auto_close=False)
p.screenshot(OUTPUT / "pv_map_model.png")
p.close()

p = pv.Plotter(off_screen=pv.OFF_SCREEN, window_size=SIZE)
p.add_mesh(rbf_mesh, scalars="score", cmap="plasma", nan_opacity=0.0,
           opacity=0.92, smooth_shading=True, scalar_bar_args=CB)
p.add_mesh(plane_mesh, color="lightgrey", opacity=0.25, show_scalar_bar=False)
p.add_mesh(cloud, color="black", point_size=12, render_points_as_spheres=True)
p.add_text("Interpolated actual surface (RBF thin-plate spline, SciPy)\n"
           "grey = theoretical plane of the model", font_size=13,
           position="upper_left")
axes_frame(p)
p.show(auto_close=False)
p.screenshot(OUTPUT / "pv_surface_rbf.png")
p.close()

p = pv.Plotter(off_screen=pv.OFF_SCREEN, window_size=SIZE)
iso_map = pv.StructuredGrid(EE, PP, np.zeros_like(EE))
iso_map["score"] = ZZ_plane.ravel(order="F")
p.add_mesh(iso_map, scalars="score", cmap="viridis", scalar_bar_args=CB)
levels = np.linspace(ZZ_plane.min(), ZZ_plane.max(), 11)[1:-1]
p.add_mesh(iso_map.contour(levels, scalars="score"), color="white",
           line_width=4, render_lines_as_tubes=True)
positions, labels = [], []
for v in levels:
    line = iso_map.contour([v], scalars="score")
    if line.n_points:
        i = int(np.argmax(line.points[:, 1]))
        positions.append(line.points[i] + np.array([0.15, -0.25, 0.1]))
        labels.append(f"{v:.1f}")
if positions:
    p.add_point_labels(np.array(positions), labels, font_size=20,
                       text_color="white", shape=None, always_visible=True)

pts = np.column_stack([agg["E"], agg["P"], np.full(len(agg), 0.05)])
p.add_point_labels(pts, [f"{int(n)}" for n in agg["n"]], font_size=18,
                   text_color="black", shape="rounded_rect",
                   shape_color="white", shape_opacity=1.0, margin=4,
                   always_visible=True)
p.show_grid(xtitle="Effectiveness E", ytitle="Versatility P", ztitle="",
            font_size=12, n_xlabels=5, n_ylabels=6, show_zlabels=False)
p.add_text("Iso-score in the (E, P) plane at median A and C\n"
           "the number in each label = countermeasures on that pair",
           font_size=13, position="upper_left")
p.view_xy()
p.camera.zoom(1.0)
p.show(auto_close=False)
p.screenshot(OUTPUT / "pv_map_iso.png")
p.close()

p = pv.Plotter(off_screen=pv.OFF_SCREEN, window_size=SIZE)
p.add_mesh(plane_mesh, color="lightgrey", opacity=0.35, show_scalar_bar=False)
p.add_mesh(stems, color="dimgrey", line_width=2)
lim = float(np.abs(df["residual"]).max())
p.add_mesh(cloud.glyph(geom=pv.Sphere(radius=0.14), scale=False, orient=False),
           scalars="residual", cmap="coolwarm", clim=(-lim, lim),
           scalar_bar_args=dict(CB, title="Residual"))
p.add_text(f"Residuals = actual score - plane (deviation due to A and C)\n"
           f"mean = {df['residual'].mean():+.2f}   "
           f"standard deviation = {df['residual'].std():.2f}   "
           f"min = {df['residual'].min():+.2f}   "
           f"max = {df['residual'].max():+.2f}",
           font_size=13, position="upper_left")
axes_frame(p)
p.show(auto_close=False)
p.screenshot(OUTPUT / "pv_residues.png")
p.close()

try:
    p = pv.Plotter(off_screen=True, window_size=SIZE)
    p.add_mesh(plane_mesh, scalars="score", cmap="viridis", opacity=0.55,
               show_scalar_bar=False)
    p.add_mesh(rbf_mesh, scalars="score", cmap="plasma", nan_opacity=0.0,
               opacity=0.75, show_scalar_bar=False)
    p.add_mesh(stems, color="dimgrey", line_width=1.5)
    p.add_mesh(cloud, scalars="score", cmap="plasma", point_size=18,
               render_points_as_spheres=True, scalar_bar_args=CB)
    axes_frame(p)
    p.export_html(OUTPUT / "quotation_interactive.html")
    p.close()
    print("HTML export: quotation_interactive.html")
except ImportError as err:
    print("HTML export skipped:", err)

print("\nFigures written to", OUTPUT.resolve())
