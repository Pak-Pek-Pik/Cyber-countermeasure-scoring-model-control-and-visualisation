from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd

matplotlib.use("Agg")

import matplotlib.pyplot as plt
from openpyxl import load_workbook

parser = argparse.ArgumentParser(
    description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
parser.add_argument("--data", type=Path, default=Path("data/parades.xlsx"),
                    help="source Excel workbook")
parser.add_argument("--output", type=Path, default=Path("figures"),
                    help="destination folder for the PNG files")
parser.add_argument("--sheet", default="Parades BdD",
                    help="name of the sheet to read")
args = parser.parse_args()

SOURCE, OUTPUT, SHEET = args.data, args.output, args.sheet
OUTPUT.mkdir(parents=True, exist_ok=True)

W_E, W_P, W_A, W_C = 0.35, 0.25, 0.20, 0.20

plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "font.size": 10,
    "axes.titlesize": 12,
    "figure.dpi": 150,
})

ws = load_workbook(SOURCE, data_only=True)[SHEET]
rows = []
for r in range(3, ws.max_row + 1):
    name = ws[f"A{r}"].value
    E, P, A, C, score = (ws[f"{c}{r}"].value for c in "DEFGH")
    if None in (E, P, score):
        continue
    rows.append({"name": name, "status": ws[f"B{r}"].value,
                 "E": E, "P": P, "A": A, "C": C, "score": score})

df = pd.DataFrame(rows)
print(f"{len(df)} countermeasures read")
print(df[["E", "P", "A", "C", "score"]].describe().round(2))

A_ref, C_ref = df["A"].median(), df["C"].median()
offset = W_A * A_ref + W_C * C_ref
print(f"\nSection at A = {A_ref} and C = {C_ref}  ->  offset = {offset:.2f}")

partial = W_E * df["E"] + W_P * df["P"] + offset
r2 = 1 - ((df["score"] - partial) ** 2).sum() / \
         ((df["score"] - df["score"].mean()) ** 2).sum()
print(f"R2 of the (E,P) plane against the actual scores: {r2:.3f}")

e_grid = np.linspace(df["E"].min(), df["E"].max(), 60)
p_grid = np.linspace(df["P"].min(), df["P"].max(), 60)
EE, PP = np.meshgrid(e_grid, p_grid)
ZZ = W_E * EE + W_P * PP + offset

fig = plt.figure(figsize=(11, 8))
ax = fig.add_subplot(111, projection="3d")

ax.plot_surface(EE, PP, ZZ, cmap="viridis", alpha=0.55,
                linewidth=0, antialiased=True, rstride=2, cstride=2)
ax.contour(EE, PP, ZZ, levels=8, zdir="z",
           offset=df["score"].min() - 0.5, cmap="viridis", linewidths=0.8)

cloud = ax.scatter(df["E"], df["P"], df["score"],
                   c=df["score"], cmap="plasma", s=45,
                   edgecolor="black", linewidth=0.4, depthshade=False)

for _, row in df.iterrows():
    z_plane = W_E * row["E"] + W_P * row["P"] + offset
    ax.plot([row["E"]] * 2, [row["P"]] * 2, [z_plane, row["score"]],
            color="grey", linewidth=0.4, alpha=0.5)

ax.set_xlabel("Effectiveness E (col. D)", labelpad=10)
ax.set_ylabel("Versatility P (col. E)", labelpad=10)
ax.set_zlabel("Score (col. H)", labelpad=8)
ax.set_title(f"Score = 0.35 E + 0.25 P + 0.20 A + 0.20 C\n"
             f"plane drawn at A = {A_ref:.0f} and C = {C_ref:.0f}, "
             f"points = {len(df)} actual countermeasures", pad=4)
ax.set_zlim(df["score"].min() - 0.6, max(ZZ.max(), df["score"].max()) + 0.2)
ax.view_init(elev=24, azim=-128)
fig.colorbar(cloud, ax=ax, shrink=0.55, pad=0.10, label="Actual score")
fig.tight_layout()
fig.savefig(OUTPUT / "quotation_surface_3D.png", bbox_inches="tight")
plt.close(fig)

fig, ax = plt.subplots(figsize=(10.5, 7.5))

im = ax.contourf(EE, PP, ZZ, levels=18, cmap="viridis", alpha=0.9)
lines = ax.contour(EE, PP, ZZ, levels=9, colors="white", linewidths=0.9)
ax.clabel(lines, inline=True, fontsize=8, fmt="%.1f")

agg = (df.groupby(["E", "P"])
         .agg(n=("name", "size"), mean_score=("score", "mean"))
         .reset_index())
ax.scatter(agg["E"], agg["P"], s=90 + 70 * agg["n"], c="white",
           edgecolor="black", linewidth=1.1, zorder=3)
for _, row in agg.iterrows():
    ax.annotate(f"{int(row['n'])}", (row["E"], row["P"]),
                ha="center", va="center", fontsize=8, fontweight="bold",
                color="black", zorder=4)

ax.set_xlabel("Effectiveness E (column D)")
ax.set_ylabel("Versatility P (column E)")
ax.set_title("Iso-score contours in the (E, P) plane\n"
             f"at A = {A_ref:.0f} and C = {C_ref:.0f}, "
             "the number in each circle = countermeasures on that pair")
ax.set_xlim(df["E"].min() - 0.6, df["E"].max() + 0.6)
ax.set_ylim(df["P"].min() - 0.6, df["P"].max() + 0.6)
fig.colorbar(im, ax=ax, label="Score")
ax.grid(alpha=0.15, linestyle=":")
fig.tight_layout()
fig.savefig(OUTPUT / "quotation_map_2D.png", bbox_inches="tight")
plt.close(fig)

fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

pivot = df.pivot_table(index="P", columns="E", values="score", aggfunc="mean")
heat = axes[0].imshow(pivot.values, origin="lower", cmap="plasma", aspect="auto")
axes[0].set_xticks(range(len(pivot.columns)), [str(c) for c in pivot.columns])
axes[0].set_yticks(range(len(pivot.index)), [str(i) for i in pivot.index])
for i in range(pivot.shape[0]):
    for j in range(pivot.shape[1]):
        v = pivot.values[i, j]
        if not np.isnan(v):
            axes[0].text(j, i, f"{v:.1f}", ha="center", va="center",
                         fontsize=7.5, color="white")
axes[0].set_xlabel("Effectiveness E")
axes[0].set_ylabel("Versatility P")
axes[0].set_title("Mean observed score per (E, P) pair")
fig.colorbar(heat, ax=axes[0], label="Score")

for p in sorted(df["P"].unique()):
    axes[1].plot(e_grid, W_E * e_grid + W_P * p + offset,
                 linewidth=1.4, label=f"P = {p}")
axes[1].scatter(df["E"], df["score"], s=18, c="black", alpha=0.45,
                zorder=3, label="actual countermeasures")
axes[1].set_xlabel("Effectiveness E")
axes[1].set_ylabel("Score")
axes[1].set_title("Sections of the plane: score(E) at fixed P")
axes[1].legend(fontsize=7, ncol=2)
axes[1].grid(alpha=0.2, linestyle=":")

fig.tight_layout()
fig.savefig(OUTPUT / "quotation_section.png", bbox_inches="tight")
plt.close(fig)

print("\n3 figures written to", OUTPUT.resolve())
